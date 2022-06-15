import matplotlib
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import logging

matplotlib.use("Agg")

dir_path = "C:/2021RnE/G4data/1~10GeV/"
load_file_path_mu = dir_path + "mu-1~10GeV" + ".xlsx"
load_file_path_e = dir_path + "e-1~10GeV" + ".xlsx"
save_fdt_dir_path = dir_path + "fdt_data/"
save_hist_dir_path = dir_path + "hist_data/"

wb_mu = load_workbook(load_file_path_mu, data_only=True)
wb_e = load_workbook(load_file_path_e, data_only=True)

wb_mu = load_workbook(load_file_path_mu, data_only=True)
wb_e = load_workbook(load_file_path_e, data_only=True)

for i in range(1, 11):  # 1~10GeV
    ws_mu = wb_mu['Sheet' + str(i)]
    ws_e = wb_e['Sheet' + str(i)]

    listCh2_mu = []
    get_cells = ws_mu['C9':'C1008']
    for row in get_cells:
        for cell in row:
            listCh2_mu.append(cell.value)
    listCh3_mu = []
    get_cells = ws_mu['D9':'D1008']
    for row in get_cells:
        for cell in row:
            listCh3_mu.append(cell.value)
    listRe_mu = []
    for j in range(1000):
        if listCh3_mu[j] == 0:
            continue
        listRe_mu.append(listCh2_mu[j] / listCh3_mu[j] * 100)

    listCh2_e = []
    get_cells = ws_e['C9':'C1008']
    for row in get_cells:
        for cell in row:
            listCh2_e.append(cell.value)
    listCh3_e = []
    get_cells = ws_e['D9':'D1008']
    for row in get_cells:
        for cell in row:
            listCh3_e.append(cell.value)
    listRe_e = []
    for j in range(1000):
        if listCh3_e[j] == 0:
            continue
        listRe_e.append(listCh2_e[j] / listCh3_e[j] * 100)

    logger = logging.getLogger("histG4data")
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler("c:/log/histG4data1_relative.log")
    fh.setLevel(logging.DEBUG)
    logger.addHandler(fh)

    bins = np.arange(0, 1000, 1)
    hist_mu, bins = np.histogram(listRe_mu, bins)
    total_number_mu = len(listRe_mu)
    hist_normal_mu = np.asarray(hist_mu) / total_number_mu
    list_hist_normal_mu = []
    for j in range(0, 999):
        list_hist_normal_mu.append(hist_normal_mu[j])

    logger.debug(list_hist_normal_mu)

    plt.hist(listRe_mu, bins=range(0, 1000, 1), density=True, alpha=1, histtype='stepfilled')
    plt.axis('off')
    plt.savefig('c:/Users/dr2mer05/Desktop/G4data/mu/mu-%dGeV.png' % i)
    plt.close()

    bins = np.arange(0, 1000, 1)
    hist_e, bins = np.histogram(listRe_e, bins)
    total_number_e = len(listRe_e)
    hist_normal_e = np.asarray(hist_e) / total_number_e
    list_hist_normal_e = []
    for j in range(0, 999):
        list_hist_normal_e.append(hist_normal_e[j])

    logger.debug(list_hist_normal_e)

    # 뮤온 히스토그램 이미지 저장
    plt.hist(listRe_mu, bins=range(0, 1000, 1), density=True, alpha=1, histtype='stepfilled')
    plt.axis('off')
    load_file_path_mu = "{}/{}%d.{}".format(save_dir_path_mu, save_file_name_mu, i, save_file_type)
    plt.savefig(load_file_path_mu)
    plt.close()
    
    plt.hist(listRe_e, bins=range(0, 1000, 1), density=True, alpha=1, histtype='stepfilled')
    plt.axis('off')
    load_file_path_e = "{}/{}%d.{}".format(save_dir_path_e, save_file_name_e, i, save_file_type)
    plt.savefig(load_file_path_e)
    plt.close()
    
