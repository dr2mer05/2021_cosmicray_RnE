import matplotlib
from openpyxl import load_workbook
import random
import numpy as np
import matplotlib.pyplot as plt
import logging

NUM = 10000
RANGE = 250

matplotlib.use("Agg")  # Agg Buffer 사용

# 초기 에너지에 대한 입자의 비율(1~10GeV)
rateOfNum = [0.199163597,
             0.14082993,
             0.114987156,
             0.099581798,
             0.089068668,
             0.081308198,
             0.075276764,
             0.070414965,
             0.066387866,
             0.062981059]

dir_path = "C:/2021RnE/G4data/1~10GeV/"
load_file_path_mu = dir_path + "mu-1~10GeV" + ".xlsx"
load_file_path_e = dir_path + "e-1~10GeV" + ".xlsx"

load_wb_mu = load_workbook(load_file_path_mu, data_only=True)
load_wb_e = load_workbook(load_file_path_e, data_only=True)

hist_sum = [0 for i in range(RANGE)]
load_wb = load_workbook(load_file_path_mu, data_only=True)

logger = logging.getLogger("fdt_mu_data")
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler(dir_path + "mu.log")
fh.setLevel(logging.DEBUG)
logger.addHandler(fh)

for j in range(1, 11):
    load_ws = load_wb_mu['Sheet'+str(j)]

    # r_e을 구해 리스트에 저장하기
    listCh2 = []
    get_cells = load_ws['C9':'C{}'.format(NUM+8)]
    for row in get_cells:
        for cell in row:
            listCh2.append(cell.value)
    listCh3 = []
    get_cells = load_ws['D9':'D{}'.format(NUM+8)]
    for row in get_cells:
        for cell in row:
            listCh3.append(cell.value)
    listRe = []
    for i in range(NUM):
        if listCh3[i] == 0:
            continue
        listRe.append(listCh2[i] / listCh3[i] * 100)

    bins = np.arange(0, RANGE, 1)
    hist, bins = np.histogram(listRe, bins)

    for i in range(0, RANGE-1):
        hist_sum[i] += hist[i] * rateOfNum[j-1]

logger.debug(hist_sum)

plt.hist(hist_sum, bins=range(0, RANGE, 1), density=True, alpha=1, histtype='stepfilled')
plt.axis('off')
plt.savefig(dir_path + "mu.png")
plt.close()


# electron
logger = logging.getLogger("fdt_e_data")
logger.setLevel(logging.DEBUG)  
fh = logging.FileHandler(dir_path + "e.log")
fh.setLevel(logging.DEBUG)
logger.addHandler(fh)

hist_sum = [0 for i in range(RANGE)]
for j in range(1, 11):
    load_ws = load_wb_e['Sheet'+str(j)]

    # r_e을 구해 리스트에 저장하기
    listCh2 = []
    get_cells = load_ws['C9':'C{}'.format(NUM+8)]
    for row in get_cells:
        for cell in row:
            listCh2.append(cell.value)
    listCh3 = []
    get_cells = load_ws['D9':'D{}'.format(NUM+8)]
    for row in get_cells:
        for cell in row:
            listCh3.append(cell.value)
    listRe = []
    for i in range(NUM):
        if listCh3[i] == 0:
            continue
        listRe.append(listCh2[i] / listCh3[i] * 100)

    bins = np.arange(0, RANGE, 1)
    hist, bins = np.histogram(listRe, bins)

    for i in range(0, RANGE-1):
        hist_sum[i] += hist[i] * rateOfNum[j-1]

logger.debug(hist_sum)

plt.hist(hist_sum, bins=range(0, RANGE, 1), density=True, alpha=1, histtype='stepfilled')
plt.axis('off')
plt.savefig(dir_path + "e.png")
plt.close()


