import matplotlib
from openpyxl import load_workbook
import random
import numpy as np
import matplotlib.pyplot as plt
import logging

NUM = 10000
RANGE = 250

matplotlib.use("Agg")  # Agg Buffer 사용

dir_path = "C:/2021RnE/G4data/1GeV/"
load_file_path_mu = dir_path + "mu-1GeV" + ".xlsx"
load_file_path_e = dir_path + "e-1GeV" + ".xlsx"

load_wb_mu = load_workbook(load_file_path_mu, data_only=True)
load_wb_e = load_workbook(load_file_path_e, data_only=True)

logger = logging.getLogger("fdt_mu_data")
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler(dir_path + "mu.log")
fh.setLevel(logging.DEBUG)
logger.addHandler(fh)

load_ws = load_wb_mu['Sheet1']

# r_e을 구해 리스트에 저장하기
listCh2 = []
get_cells = load_ws['C9':'C{}'.format(NUM+8)]
for row in get_cells:
    for cell in row:
        listCh2.append(cell.value)
listCh3 = []
get_cells = load_ws['D9':'D{}'.format(NUM+8)]
for row in get_cells:
    for cell in row:
        listCh3.append(cell.value)
listRe = []
for i in range(NUM):
    if listCh3[i] == 0:
        continue
    listRe.append(listCh2[i] / listCh3[i] * 100)

bins = np.arange(0, RANGE, 1)
hist, bins = np.histogram(listRe, bins)
hist_log = [0 for i in range(RANGE)]
for i in range(0, RANGE-1):
    hist_log[i] += hist[i]
logger.debug(hist_log)


plt.hist(hist, bins=range(0, RANGE, 1), density=True, alpha=1, histtype='stepfilled')
plt.axis('off')
plt.savefig(dir_path + "mu.png")
plt.close()


# electron
logger = logging.getLogger("fdt_e_data")
logger.setLevel(logging.DEBUG)  
fh = logging.FileHandler(dir_path + "e.log")
fh.setLevel(logging.DEBUG)
logger.addHandler(fh)

load_ws = load_wb_e['Sheet1']

# r_e을 구해 리스트에 저장하기
listCh2 = []
get_cells = load_ws['C9':'C{}'.format(NUM+8)]
for row in get_cells:
    for cell in row:
        listCh2.append(cell.value)
listCh3 = []
get_cells = load_ws['D9':'D{}'.format(NUM+8)]
for row in get_cells:
    for cell in row:
        listCh3.append(cell.value)
listRe = []
for i in range(NUM):
    if listCh3[i] == 0:
        continue
    listRe.append(listCh2[i] / listCh3[i] * 100)

bins = np.arange(0, RANGE, 1)
hist, bins = np.histogram(listRe, bins)
hist_log = [0 for i in range(RANGE)]
for i in range(0, RANGE-1):
    hist_log[i] += hist[i]
logger.debug(hist_log)

plt.hist(hist, bins=range(0, RANGE, 1), density=True, alpha=1, histtype='stepfilled')
plt.axis('off')
plt.savefig(dir_path + "e.png")
plt.close()


